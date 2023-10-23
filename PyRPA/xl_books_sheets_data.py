import openpyxl as xl

# 새 워크북생성: xl.Workbook()
# 기존 워크북읽기: xl.load_workbook('경로명/파일명.xlsx')
# 기존 워크북읽기(엑셀 수식 계산완료후): 
#      xl.load_workbook('경로명/파일명.xlsx',data_only=True)
# 워크북 닫기: xl.close()

# 워크북에서 워크시트객체 선택
#   워크북객체.active
#   워크북객체.worksheets[시트번호] 
#     시트번호: 0에서부터 시작 (인덱스)
#   워크시트이름 활용 
#   워크북객체[시트이름]
#   book['Sheet1']
#   워크북객체.sheetnames: 워크북에 있는 워크시트이름들을 반환

# 워크북에 시트생성
#    워크북객체.create_sheet(title='시트명')

# 워크북에 시트 복사
#    복사되는 워크시트객체 = 워크북객체.copy_worksheet(복사할워크시트객체)

# 워크시트에 이름변경: 워크시트객체.title = '새로운워크시트명'

# 워크시트의 삭제: 워크북객체.remove(삭제할워크시트객체)


#import openpyxl as xl
import random

# 당첨 시트 번호 정하기 --- (1)
win_num = random.randint(1,100)

# 새 워크북 생성 --- (2)
book = xl.Workbook()
book.active["B2"] = "'당첨'이라고 적힌 시트를 찾아보자"

# 반복문을 돌며 100개의 시트 만들기 --- (3)
for i in range(1,101):
    # 새 워크시트 생성 --- (4)
    sname = str(i) + "번"
    sheet = book.create_sheet(title=sname)
    # 시트에 쓸 단어 정하기
    word = "꽝"
    if i == win_num: word = "당첨"
    # 임팩트가 있도록 화면을 word로 채우기 --- (5)
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y+1, x+1)
            c.value = word

# 파일 저장 --- (6)
book.save("./output/sheet_game.xlsx")
print("ok, winning number=", win_num)