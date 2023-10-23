import openpyxl as xl

infile = './input/name2.xlsx'
outfile = './output/name_combine.xlsx'

inxlfile = xl.load_workbook(infile)
inSheet = inxlfile.worksheets[0] # active sheet 선택과 비교

outxlfile = xl.Workbook()
outSheet = outxlfile.active

for row in inSheet.iter_rows():
    sung  = row[0].value
    myung = row[1].value

    name = sung + ' '+myung
    outSheet.append([name])

outxlfile.save(outfile)    
