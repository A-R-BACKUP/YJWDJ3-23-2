import openpyxl as xl

infile = './input/name1.xlsx'
outfile = './output/name_split.xlsx'

inxlfile = xl.load_workbook(infile)
inSheet = inxlfile.worksheets[0]

outxlfile = xl.Workbook()
outSheet = outxlfile.active

for row in inSheet.iter_rows(min_row=1):
    name = row[0].value.strip() # 좌우 공백문자열 삭제
    if ' ' in name:
        sung, myung = name.split(' ') # 리스트의 구조분해할당
    else:
        sung, myung = name[0], name[1:] # 문자열의 slicing

    outSheet.append([sung, myung])
outxlfile.save(outfile)
