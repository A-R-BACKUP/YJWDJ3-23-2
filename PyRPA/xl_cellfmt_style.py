import openpyxl as xl

book = xl.Workbook()
sheet = book.active

# 셀 너비와 높이 조정
sheet.column_dimensions['b'].width= 40
sheet.row_dimensions[2].height= 40

cell = sheet['b2']
cell.value = '微笑みは幸せの種です。'

from openpyxl.styles.alignment import Alignment
cell.alignment = Alignment(
    horizontal='center',
    vertical='center'
)

from openpyxl.styles.borders import Border,Side
cell.border = Border(
    top= Side(style='thin',color='000000'),
    right=Side(style='thick',color='FF0000'),
    bottom=Side(style='medium',color='00FF00'),
    left=Side(style='double',color='0000FF')
)

from openpyxl.styles import Font
cell.font = Font(
    size = 14,
    bold=True,
    italic=True,
    color='AAFFBB'
)

from openpyxl.styles import PatternFill
cell.fill = PatternFill(
    fill_type='solid',
    fgColor='DD2255'
)

book.save('output/cellfmt_styles.xlsx')