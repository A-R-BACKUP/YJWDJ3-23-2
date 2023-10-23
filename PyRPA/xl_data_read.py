import openpyxl as xl

# 엑셀에서 데이터 읽기
# - 셀객체.value
# - 셀객체의 지정방법
# 1. 셀주소활용
#    sheet['셀주소']: 셀객체
#    sheet['셀주소'].value
#    셀주소: 열 - A,B,...,AA,AB...
#            행 - 1, 2,... 
# 2. 셀함수활용
#    sheet.cell(행번호,열번호): 셀객체
#    sheet.cell(행번호,열번호).value
#    행번호, 열번호: 1,2,...

book = xl.load_workbook('output/write_cellname.xlsx')

sheet = book.active

""" print(sheet['A1'].value)

print(sheet.cell(1,1).value) """

""" 
# 일정 범위의 엑셀데이터 읽기 
1. for 문활용

2. 왼쪽위,오른쪽아래 지정
   left top, right bottom  

3. iter_rows(), iter_cols()
"""

# 1. for 활용법

""" for y in range(2,5): # 2~4
    r = []
    for x in range(2,5):
        v = sheet.cell(y,x).value
        r.append(v)
    print(r)
     """

# 2. 주소지정 - left top:right bottom
# 1) sheet['왼쪽위셀주소':'오른쪽아래셀주소']
# 2) sheet['왼쪽위셀주소:오른쪽아래셀주소']

#print(sheet['B2':'D4'])
""" print(sheet['B2:D4'])

for row in sheet['B2:D4']:
    r= []
    for cell in row:
        r.append(cell.value)
    print(r)
 """

 # 컴프리헨션 (comprehension) 문법: 리스트 등 반복가능한 데이터일때
""" for row in sheet['B2:D4']:
    print([c.value for c in row])

 """
# 3. iter_rows(), iter_cols()
# iter_rows(min_row,max_row,min_col,max_col)
#           최소행, 최대행,  최소열, 최대열의 인덱스값
# 반환값: iterator 객체 - next()

""" it = sheet.iter_rows(min_row=2,max_row=4,min_col=2,max_col=4)

for row in it:
    r=[]
    for cell in row:
        r.append(cell.value)
    print(r) """

# 셀주소, 행열인덱스 상호 변환
# 셀객체.row, 셀객체.column
# 셀객체.coordinate 

cell = sheet['C2']
(row,col) = (cell.row,cell.column)
print('C2 = ({0},{1})'.format(row,col))

cell= sheet.cell(row=2,column=3)
cdt = cell.coordinate # 셀주소값
print('(2,3) = {}'.format(cdt))


