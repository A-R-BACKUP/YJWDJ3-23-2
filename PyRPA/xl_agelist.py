import openpyxl as xl
import datetime as dt
import os


dir = 'output'
if dir not in os.listdir():
    os.mkdir(dir)

book = xl.Workbook()
sheet = book.active    

thisyear = dt.datetime.now().year #  실행할때의 년도 구하기
sheet['A1']='출생년도'
sheet['B1']='우리나라나이'
sheet['C1']='만나이(생일 후)'
sheet['D1']='만나이(생일 전)'

sheet.column_dimensions['B'].width = 12
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15

sheet.row_dimensions[1].height = 40
#sheet.row_dimensions['1'].height = 20 ==>에러

for i in range(80): # 0~79
    birthyear = thisyear-i

    korea_age = thisyear - birthyear + 1
    man_age = {'after': korea_age-1, 'before': korea_age-2} # dictionary
    

    year_cell = sheet.cell(i+2,1)
    year_cell.value = str(birthyear)+"년생"
    
    age_cell = sheet.cell(i+2,2)
    age_cell.value = str(korea_age)+"세"

    age_cell = sheet.cell(i+2,3)
    age_cell.value = '만'+str(man_age['after'])+"세"

    age_cell = sheet.cell(i+2,4)
    age_cell.value = '만'+str(man_age['before'])+"세"

sheet['D2']='-'

book.save(dir+'/years.xlsx')



