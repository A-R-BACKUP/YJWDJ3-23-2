import openpyxl as excel

""" book = excel.Workbook() # 새 워크북 생성 : 통합문서 생성, 오픈

sheet = book.active # 현재 활성 워크시트 반환

sheet['A1']= '안녕하세요? 일본IT과 입니다.'  #셀에 데이터 입력

sheet['C4']= '열심히합시다.' #셀에 데이터 입력

book.save('./openpyxl.xlsx') # 파일저장
 """

""" book = excel.load_workbook('./openpyxl.xlsx') # 기존 워크북 오픈

sheet = book.worksheets[0] # 첫번째 워크시트 가져오기

cell1 = sheet['A1']
cell2 = sheet['c4']

print(cell1.value)
print(cell2.value) """

# openpyxl 임포트
# import openpyxl [as 별명]
# 워크북 생성, 오픈
# - Workbook(), load_workbook()
# 워크북에서 워크시트를 가져오기
# - 워크북명.active
# - 워크북명.worksheets[인덱스]
# - 워크북명.worksheets['시트명']
# 셀에 읽기/쓰기
# - 워크시트명['셀주소'].value
# - 워크시트명['셀주소'] = '문자열'
# 셀주소
# - xy
# - x: A~AA~AZ~BA~
# - y: 1~
# - 셀명.coordinate # 특정셀의 셀주소(좌표)알아내기
# 워크북의 저장
# 워크북명.save(경로문자열)

# range(n)
# - 0~n-1
# range(n,m)
# - n~m-1
# range(n,m,s)
# - n,n+s,n+s+s,....m-1


