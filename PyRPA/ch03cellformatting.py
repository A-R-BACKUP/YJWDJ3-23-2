""" import openpyxl as xl
book = xl.load_workbook('./input/stock-data.xlsx')

sheet = book.active

print(sheet['A4'].number_format)

print(sheet['A4'].value) # datetime.datetime()
print(type(sheet['A4'].value)) """

import openpyxl as xl
import datetime as dt

infile = './input/stock-data.xlsx'
outfile = './output/stock-data-formatting.xlsx'
wantedFormat = 'yy\/mm\/dd'

def change_format(inf,outf):
    book = xl.load_workbook(inf)
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                ch_cell_format(cell)

    book.save(outf)

def ch_cell_format(cell):
    if type(cell.value) is dt.datetime:
        cell.number_format = wantedFormat

if __name__ == '__main__':
    change_format(infile,outfile)
