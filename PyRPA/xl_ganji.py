import openpyxl as xl
import datetime as dt
import os

gan = ['갑을병정무기경신임계','甲乙丙丁戊己庚辛壬癸','청청적적황황백백흑흑']
ji = ['자축인묘진사오미신유술해','子丑寅卯辰巳午未申酉戌亥',['쥐','소','호랑이','토끼','용','뱀','말','양','원숭이','닭','개','돼지']]

def year_to_ganji(year):
    i = (year-4) % 10
    j = (year-4) % 12

    ganji = gan[0][i]+ji[0][j]+'('+ gan[1][i]+ji[1][j] +')'

    color_animal = gan[2][i]+'색 '+ ji[2][j]
    
    return ganji, color_animal


dir = 'output'
if dir not in os.listdir():
    os.mkdir(dir)

book = xl.Workbook()
sheet = book.active   

sheet['A1']='연도'
sheet['B1']='간지'
sheet['c1']='동물'

start_y = now = dt.datetime.now().year

for i in range(100):
    year = start_y -i   
    result = year_to_ganji(year) # list

    ganji = result[0]
    col_ani = result[1]

    sheet.cell((2+i),1, str(year)+'년')
    sheet.cell((2+i),2, ganji)
    sheet.cell((2+i),3, col_ani)
    print(year,'=',ganji,',',col_ani)

book.save(dir+'/ganji.xlsx')    

