import openpyxl as xl

book = xl.load_workbook('input/all-customer.xlsx')

sheet = book['명부']

customers = [['이름','주소','구매 플랜']]

for row in sheet.iter_rows(min_row=3):
    values = [v.value for v in row]
    if values[0] is None: break

    area =values[1]
    if area == '서울' or area == '경기' or area == '인천':
        customers.append(values)
        print(values)

new_book = xl.Workbook()
new_sheet = new_book.active
new_sheet['A1'] = '수도권 고객 명단'

for row, row_val in enumerate(customers):
    for col, col_val in enumerate(row_val):
        cell = new_sheet.cell(row+2,col+1)
        cell.value = col_val

new_book.save('output/sheet_area.xlsx')


