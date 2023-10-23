import openpyxl as xl
import os

book = xl.Workbook()
sheet = book.active

# 데이터의 쓰기
# 1. dictionary사용형태: 셀주소 사용
# 워크시트명['셀주소'] = '저장할 데이터'
sheet['A1'] = '쓰기연습'

# 2. 메소드 사용형태: cell(row,column,value)
# 1) 메소드 호출
# 워크시트명.cell(row=행번호, column=열번호, value='저장할 데이터')
# 셀주소: A2
sheet.cell(row=2,column=1,value='하늘은 스스로 돕는 자를 돕는다')

# 2) 메소드 호출+속성
# 셀명 = 워크시트명.cell(row=행번호, column=열번호)
# 셀명.value = '저장할 데이터'
# 셀주소: A3
셀명 = sheet.cell(row=3,column=1)
셀명.value = '낙숫물이 바위를 뚫는다'

sheet.cell(row=4,column=1).value = '이거 가능할까?'

dir = 'output'
if dir not in os.listdir():
    os.mkdir(dir)


book.save('./output/write_cell.xlsx')


