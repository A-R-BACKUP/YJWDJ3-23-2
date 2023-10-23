
import openpyxl as xl
import datetime as dt
from openpyxl.styles import PatternFill

infile = './input/stock-data.xlsx'
outfile = './output/stock-data-formatting-color.xlsx'
#wantedFormat = 'yy\/mm\/dd'
limit = dt.datetime(2020,1,1)

def change_format_color(inf,outf):
    book = xl.load_workbook(inf)
    for sheet in book.worksheets:
        for row in sheet.iter_rows(min_row=4):
            ch_cell_format(row)

    book.save(outf)

def ch_cell_format(row):
    date = row[0].value
    if type(date) is not dt.datetime:
        return
    if(date<limit):
        red = PatternFill(
            fill_type ='solid',
            fgColor='ff0000'
        )
        for cell in row:
            cell.fill =red

if __name__ == '__main__':
    change_format_color(infile,outfile)