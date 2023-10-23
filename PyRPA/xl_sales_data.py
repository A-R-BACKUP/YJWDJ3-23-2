import openpyxl as xl

#book = xl.load_workbook('./input/monthly_sales.xlsx')
book = xl.load_workbook(
            './input/monthly_sales.xlsx', # 읽을 엑셀파일
            data_only=True # 엑셀 수식을 실행해서 데이터만 읽어오기 옵션
        )
sheet = book.active

""" rows = sheet['A3:F9']

for row in rows:
    values = [cell.value for cell in row]
    print(values) """

# 데이터의 갯수를 명확이 모를 경우 전체 데이터 읽기
# 워크시트객체.max_row, 워크시트객체.max_column

# print(sheet.max_row, sheet.max_column)
# 문제점: 셀의 서식만 있고 실제 데이터가 없는 경우도 
# 워크시트객체.max_row, 워크시트객체.max_column값에 영향
# 해결책1: max_column또는max_row까지 반복하면서 데이터가 있는지 
# 조사하여 데이터가 없으면 None: 반복문 break
# 해결책2: iter_rows()에서 max_row인수 생략하면 데이터 있는 셀까지만 반복처리

""" rows = sheet['A3:F999']
for row in rows:
    values = [cell.value for cell in row]
    if values[0] is None: break
    print(values) """

for row in sheet.iter_rows(min_row=3):
    values = [cell.value for cell in row]
    if values[0] is None: break
    print(values)


