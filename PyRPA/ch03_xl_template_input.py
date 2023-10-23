import openpyxl as xl

template_file = './input/invoice-template.xlsx'
save_file = './output/invoice-sample.xlsx'

name = '노영환'
subject = '1월 청구분'
items = [
    ['사과',5,3000],
    ['바나나',8,1200],
    ['메론',1,12000]
]

book = xl.load_workbook(template_file)
sheet = book.active

sheet['B4'] = name
sheet['C10'] = subject

total = 0
for i, item in enumerate(items):
    summary, count, price = item # 구조분해 할당
    subtotal = count * price
    total+=subtotal

    row = 15+i
    sheet.cell(row,2, summary)
    sheet.cell(row,5, count)
    sheet.cell(row,6, price)
    sheet.cell(row,7, subtotal)
    
sheet['C11'] = total

book.save(save_file)

    

