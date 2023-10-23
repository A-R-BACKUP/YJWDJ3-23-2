import openpyxl as xl

book = xl.load_workbook('input/all-customer.xlsx')

sheet = book['명부']

for row in sheet.iter_rows(min_row=3):
    values = [v.value for v in row]
    if values[0] is None: break

    print(values)

    (name,area,plan)=values
    sheet_name = plan+'플랜'

    if sheet_name not in book.sheetnames: 
        to_sheet = book.create_sheet(title=sheet_name)
        to_sheet.append(['이름','주소','구매 플랜'])
    else:
        to_sheet = book[sheet_name]
    
    to_sheet.append(values)


book.save('output/sheet_plans.xlsx')